import os
import pandas as pd
import requests
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET
import logging
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Font

# Отключение предупреждений о небезопасных запросах
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Настройка логирования
logging.basicConfig(filename='logs.txt', level=logging.INFO, 
                    format='%(asctime)s %(levelname)s:%(message)s', filemode='w')

# Путь к файлу с ключевыми запросами
queries_file = 'queries.txt'
# Путь для сохранения результатов
output_file = 'parsed_data.xlsx'
# Данные для API
user = ' '
key = ' '
api_url = f'https://xmlstock.com/yandex/xml/?user={user}&key={key}&query='

# Функция для парсинга выдачи Яндекс через API
def search_yandex(query):
    try:
        url = api_url + requests.utils.quote(query)
        response = requests.get(url)
        response.raise_for_status()
        return response.text
    except requests.exceptions.RequestException as e:
        logging.error(f"Error during API request for query '{query}': {e}")
        return None

# Функция для получения топовых URL из XML ответа
def get_top_urls(query):
    xml_response = search_yandex(query)
    if xml_response is None:
        return []
    
    root = ET.fromstring(xml_response)
    if root.find(".//error") is not None:
        error_code = root.find(".//error").get("code")
        if error_code == "15":  # код ошибки капчи может быть другим, это пример
            logging.error(f"Captcha encountered for query '{query}'")
        else:
            logging.error(f"Error in XML response for query '{query}': {root.find('.//error').text}")
        return []

    urls = []
    for doc in root.findall(".//doc"):
        url = doc.find('url').text
        urls.append(url)
    return urls[:10]  # Возвращаем топ 10 URL

# Функция для получения H1, H2-H6 и Title
def get_page_data(url):
    try:
        response = requests.get(url, timeout=10, verify=False)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')

        # Поиск H1
        h1 = soup.find('h1')
        if h1 is None:
            h1 = soup.find('h1', class_=True)
        h1_text = h1.get_text(strip=True) if h1 else ''

        # Поиск H2-H6
        headers = []
        for i in range(2, 7):
            header = soup.find_all(f'h{i}')
            headers.extend([h.get_text(strip=True) for h in header])
        
        # Поиск Title
        title = soup.title.string.strip() if soup.title else ''
        
        return h1_text, headers, title
    except Exception as e:
        logging.error(f"Error processing URL '{url}': {e}")
        return '', [], ''

# Загрузка ключевых запросов
with open(queries_file, 'r', encoding='utf-8') as f:
    queries = [line.strip() for line in f if line.strip()]

# Создание словаря для хранения уникальных URL и соответствующих запросов
url_query_mapping = {}

# Сбор данных
for query in queries:
    logging.info(f"Processing query: {query}")
    top_urls = get_top_urls(query)
    for url in top_urls:
        if url not in url_query_mapping:
            url_query_mapping[url] = [query]
        else:
            url_query_mapping[url].append(query)

# Создание списка для хранения результатов
results = []

# Сбор данных для каждого уникального URL
for url, query_list in url_query_mapping.items():
    logging.info(f"Processing URL: {url}")
    h1, headers, title = get_page_data(url)
    results.append((url, '\n'.join(query_list), h1, '\n'.join(headers), title))

# Конвертация результатов в DataFrame
results_df = pd.DataFrame(results, columns=['URL', 'Ключевые запросы', 'H1', 'H2-H6', 'Title'])

# Сохранение результатов в файл Excel
output_file_path = os.path.join(os.path.dirname(__file__), output_file)

# Создаем Excel файл и записываем данные
wb = Workbook()
ws = wb.active
ws.title = "Parsed Data"

# Заголовки
headers = ['URL', 'Ключевые запросы', 'H1', 'H2-H6', 'Title']
ws.append(headers)

# Запись данных
for row in results:
    ws.append(row)

# Выделение жирным строк с наибольшим количеством ключей
max_queries_count = max(len(url_query_mapping[url]) for url in url_query_mapping)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    cell = row[0]
    query_count = len(cell.value.split('\n'))
    if query_count == max_queries_count:
        for cell in ws[cell.row]:
            cell.font = Font(bold=True)

# Сохранение файла
wb.save(output_file_path)

logging.info(f"Data has been parsed and saved to {output_file_path}")
print(f"Data has been parsed and saved to {output_file_path}")
